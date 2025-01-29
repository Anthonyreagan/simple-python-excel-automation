import openpyxl as xl

wb = xl.load_workbook("transaction.xlsx")
sheet = wb['Sheet1']
cell = sheet.cell(1,1)
print(sheet.max_row)

for row in range (2, sheet.max_row +1 ):
    cell = sheet.cell(row , 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 6)
    corrected_price_cell.value = corrected_price


wb.save('transaction2.xlsx')